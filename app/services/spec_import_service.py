from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

ALLOWED_PART_TYPES = {"деталь", "деталь_св", "деталь_кон"}

ROUTES_BY_TYPE = {
    "деталь": ["laser", "bend"],
    "деталь_св": ["laser", "bend", "weld"],
    "деталь_кон": ["laser", "bend"],
}


@dataclass
class ImportedItem:
    part_number: str
    name: str
    metal: str | None
    thickness: float | None
    qty_per_product: int
    part_type: str


@dataclass
class ImportSummary:
    created_items: int
    created_routes: int


def _to_int(value: object, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip().replace(",", ".")
    if not text:
        return default

    try:
        return int(float(text))
    except ValueError:
        return default


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", ".")
    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def _to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_spec_excel(content: bytes) -> list[ImportedItem]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    worksheet = workbook["Лист1"]

    rows = worksheet.iter_rows(min_row=2, values_only=True)
    parsed: list[ImportedItem] = []

    for row in rows:
        part_number = _to_str(row[1] if len(row) > 1 else None)
        name = _to_str(row[2] if len(row) > 2 else None)
        qty_per_product = _to_int(row[3] if len(row) > 3 else None)
        part_type_raw = _to_str(row[5] if len(row) > 5 else None)
        thickness = _to_float(row[6] if len(row) > 6 else None)
        metal = _to_str(row[7] if len(row) > 7 else None) or None

        part_type = part_type_raw.lower()
        if not part_number:
            continue
        if part_type not in ALLOWED_PART_TYPES:
            continue

        parsed.append(
            ImportedItem(
                part_number=part_number,
                name=name,
                metal=metal,
                thickness=thickness,
                qty_per_product=qty_per_product,
                part_type=part_type,
            )
        )

    return parsed


def import_specification(
    connection: sqlite3.Connection,
    *,
    type_id: int,
    type_quantity_plan: int,
    items: list[ImportedItem],
) -> ImportSummary:
    connection.execute(
        """
        DELETE FROM routes
        WHERE item_id IN (SELECT id FROM items WHERE type_id = ?)
        """,
        (type_id,),
    )
    connection.execute("DELETE FROM items WHERE type_id = ?", (type_id,))

    created_routes = 0
    for item in items:
        total_qty = item.qty_per_product * type_quantity_plan
        cursor = connection.execute(
            """
            INSERT INTO items (
                type_id,
                part_number,
                name,
                metal,
                thickness,
                qty_per_product,
                total_qty
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                type_id,
                item.part_number,
                item.name,
                item.metal,
                item.thickness,
                item.qty_per_product,
                total_qty,
            ),
        )
        item_id = cursor.lastrowid

        for index, stage_name in enumerate(ROUTES_BY_TYPE[item.part_type], start=1):
            connection.execute(
                """
                INSERT INTO routes (item_id, stage_name, order_index)
                VALUES (?, ?, ?)
                """,
                (item_id, stage_name, index),
            )
            created_routes += 1

    return ImportSummary(created_items=len(items), created_routes=created_routes)
